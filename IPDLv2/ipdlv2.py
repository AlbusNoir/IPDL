#!/usr/bin/env python
"""
IPD_LOOKUP TOOL
Utilizes PyDNSBL to determine if an IP or domain is malicious or not
"""

import pydnsbl
from pydnsbl.providers import BASE_PROVIDERS
from pydnsbl.providers import Provider

import openpyxl

__author__ = "Kaleb Sego"
__copyright__ = "Copyright 2021"
__license__ = "GPLv3"
__version__ = "2.1.0"
__maintainer__ = "Kaleb Sego"
__contact__ = "https://github.com/AlbusNoir"

wb = openpyxl.load_workbook('ipdl_output.xlsx')
ipsheet = wb['IP']
domsheet = wb['DOMAIN']

# pull in all providers, combining providers.txt and base_providers
plist = []

with open('providers.txt') as f:
    for line in f:
        plist.append(line.rstrip())

more_providers = list(map(Provider, plist))

providers = BASE_PROVIDERS + more_providers
ip_checker = pydnsbl.DNSBLIpChecker(providers = providers)

domain_checker = pydnsbl.DNSBLDomainChecker()


def main():
    """Main driver function"""

    ip_list = []
    dom_list = []

    # only grab col A rows 2:x and add to ip_list, ignore None
    for row_cells in ipsheet.iter_rows(min_row=2, max_col=1):
        for cell in row_cells:
            if cell.value is not None:
                ip_list.append(cell.value)

    # only grab col B rows 2:x and add to dom_list, ignore None
    for row_cells in domsheet.iter_rows(min_row=2,max_col=1):
        for cell in row_cells:
            if cell.value is not None:
                dom_list.append(cell.value)


    check_input(ip_list, dom_list)


def check_input(iplist, domlist):
    ips = iplist
    doms = domlist


    ip_output = []
    dom_output = []

    for ip in ips:
        result = ip_checker.check(ip)
        bl = result.blacklisted

        if bl == 'True':
            ip_output.append('Blacklisted')
        else:
            ip_output.append('Not Blacklisted')

    for dom in doms:
        result = domain_checker.check(dom)
        bl = result.blacklisted

        if bl == 'True':
            dom_output.append('Blacklisted')
        else:
            dom_output.append('Not Blacklisted')

    # idk the output isn't working so
    '''
    for row_cells in ipsheet.iter_rows(min_row=2, min_col=2):
        for cell in range(len(ip_output)):
            for idx,ip in enumerate(ip_output, len(ip_output)):
                cell.value = ip
'''

if __name__ == '__main__':
    main()
